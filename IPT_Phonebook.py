import sqlite3
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import tkinter.messagebox as messagebox 
import openpyxl

def create_table():
    c.execute('''
        CREATE TABLE IF NOT EXISTS App_DB (
            id INTEGER PRIMARY KEY,
            column1 TEXT,
            column2 TEXT,
            column3 TEXT,
            column4 TEXT,
            column5 TEXT
        )
    ''')
    conn.commit()

# 데이터베이스 연결 및 테이블 생성
conn = sqlite3.connect('data.db')
c = conn.cursor()
create_table()

def insert_data_from_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            c.execute("INSERT INTO App_DB (column1, column2, column3, column4, column5) VALUES (?, ?, ?, ?, ?)", row)
        conn.commit()
        update_treeview()

def export_data_to_excel():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['기계선번', '내선번호', '수평선번', '사용자명', '기타입력'])
        
        c.execute("SELECT * FROM App_DB")
        rows = c.fetchall()
        for row in rows:
            ws.append(row[1:])
        
        wb.save(file_path)

def delete_selected_row(event=None):
    selected_item = tree.selection()
    if selected_item:
        confirmation = messagebox.askyesno("삭제 확인", "선택한 행을 삭제하시겠습니까?")
        if confirmation:
            for item in selected_item:
                values = tree.item(item, "values")
                c.execute("DELETE FROM App_DB WHERE id = ?", (values[0],))
            conn.commit()
            update_treeview()
    else:
        messagebox.showwarning("선택 오류", "삭제할 행을 선택해주세요.")

def focus_next_entry(event):
    event.widget.tk_focusNext().focus()
    return "break"

def insert_data(event):
    data = (
        entry_column1.get(),
        entry_column2.get(),
        entry_column3.get(),
        entry_column4.get(),
        entry_column5.get()
    )
    
    response = messagebox.askquestion("저장 확인", "정말로 저장하시겠습니까?")
    if response == "yes":
        c.execute("INSERT INTO App_DB (column1, column2, column3, column4, column5) VALUES (?, ?, ?, ?, ?)", data)
        conn.commit()
        clear_entries()
        update_treeview()

def clear_entries():
    entry_column1.delete(0, tk.END)
    entry_column2.delete(0, tk.END)
    entry_column3.delete(0, tk.END)
    entry_column4.delete(0, tk.END)
    entry_column5.delete(0, tk.END)

def update_treeview():
    tree.delete(*tree.get_children())
    c.execute("SELECT * FROM App_DB")
    rows = c.fetchall()

    # 내선번호를 기준으로 오름차순으로 정렬
    sorted_rows = sorted(rows, key=lambda row: int(row[2]))  # 여기서 2는 내선번호의 인덱스

    for row in sorted_rows:
        tree.insert('', 'end', values=row)

# 수정 팝업 창을 만드는 함수
def open_edit_popup(selected_item):
    global edit_popup
    edit_popup = tk.Toplevel(app)
    edit_popup.title("수정")

    # 화면 중앙에 위치하도록 설정
    window_width = 205  # 팝업 창의 폭
    window_height = 135  # 팝업 창의 높이

    # 화면 크기 가져오기
    screen_width = app.winfo_screenwidth()
    screen_height = app.winfo_screenheight()

    # 화면 중앙에 위치하도록 x, y 좌표 계산
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2

    # 팝업 창을 화면 중앙에 배치
    edit_popup.geometry(f"{window_width}x{window_height}+{x}+{y}")

    # 프레임 생성
    edit_frame = tk.Frame(edit_popup)
    edit_frame.pack(expand=True, fill='both')  # 프레임을 Toplevel에 채우도록 설정

    # 선택된 항목의 데이터 가져오기
    values = tree.item(selected_item, "values")
    id_value = values[0]
    column1_value = values[1]
    column2_value = values[2]
    column3_value = values[3]
    column4_value = values[4]
    column5_value = values[5]

    # 수정 팝업 창에 엔트리 위젯 추가
    label_column1 = tk.Label(edit_frame, text="기계선번")
    entry_column1 = tk.Entry(edit_frame)
    label_column2 = tk.Label(edit_frame, text="내선번호")
    entry_column2 = tk.Entry(edit_frame)
    label_column3 = tk.Label(edit_frame, text="수평선번")
    entry_column3 = tk.Entry(edit_frame)
    label_column4 = tk.Label(edit_frame, text="사용자명")
    entry_column4 = tk.Entry(edit_frame)
    label_column5 = tk.Label(edit_frame, text="기타입력")
    entry_column5 = tk.Entry(edit_frame)

    label_column1.grid(row=0, column=0)
    entry_column1.grid(row=0, column=1)
    label_column2.grid(row=1, column=0)
    entry_column2.grid(row=1, column=1)
    label_column3.grid(row=2, column=0)
    entry_column3.grid(row=2, column=1)
    label_column4.grid(row=3, column=0)
    entry_column4.grid(row=3, column=1)
    label_column5.grid(row=4, column=0)
    entry_column5.grid(row=4, column=1)

    # 기존 데이터로 엔트리 위젯 초기화
    entry_column1.insert(0, column1_value)
    entry_column2.insert(0, column2_value)
    entry_column3.insert(0, column3_value)
    entry_column4.insert(0, column4_value)
    entry_column5.insert(0, column5_value)

    # 수정된 데이터를 저장하는 함수
    def save_edited_data():
        updated_data = (
            entry_column1.get(),
            entry_column2.get(),
            entry_column3.get(),
            entry_column4.get(),
            entry_column5.get()
        )

        response = messagebox.askquestion("저장 확인", "수정된 내용을 저장하시겠습니까?")
        if response == "yes":
            # 데이터베이스 업데이트
            c.execute("UPDATE App_DB SET column1=?, column2=?, column3=?, column4=?, column5=? WHERE id=?", (*updated_data, id_value))
            conn.commit()
            edit_popup.destroy()
            update_treeview()

    # 저장 버튼 추가
    save_button = tk.Button(edit_frame, text="저장", command=save_edited_data)
    save_button.grid(row=5, column=0, columnspan=2)
    
    entry_column1.focus()
    entry_column1.bind("<Return>", focus_next_entry)
    entry_column2.bind("<Return>", focus_next_entry)
    entry_column3.bind("<Return>", focus_next_entry)
    entry_column4.bind("<Return>", focus_next_entry)
    # 엔트리 위젯에서 엔터를 눌렀을 때 저장 기능 추가
    entry_column5.bind("<Return>", lambda event: save_edited_data())
    # ESC 키 바인딩하여 팝업 창 닫기
    edit_popup.bind("<Escape>", lambda event: edit_popup.destroy())

conn = sqlite3.connect('data.db')
c = conn.cursor()

# 윈도우 생성
app = tk.Tk()
app.title("선번장 (문의 : 김항우과장)")

# Consolas 폰트 설정
app.option_add("*Font", "Consolas 10")

# 창의 크기 설정
window_width = 1024
window_height = 730
# 모니터 화면 크기 가져오기
screen_width = app.winfo_screenwidth()
screen_height = app.winfo_screenheight()
# 창을 화면 중앙에 위치시키기 위한 x, y 좌표 계산
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
# 창을 화면 중앙에 배치
app.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 프레임 1: 엔트리와 라벨
entry_frame = tk.Frame(app)
entry_frame.pack(pady=5)

label_column1 = tk.Label(entry_frame, text="기계선번")
entry_column1 = tk.Entry(entry_frame)
label_column2 = tk.Label(entry_frame, text="내선번호")
entry_column2 = tk.Entry(entry_frame)
label_column3 = tk.Label(entry_frame, text="수평선번")
entry_column3 = tk.Entry(entry_frame)
label_column4 = tk.Label(entry_frame, text="사용자명")
entry_column4 = tk.Entry(entry_frame)
label_column5 = tk.Label(entry_frame, text="기타입력")
entry_column5 = tk.Entry(entry_frame)

label_column1.pack(side='left')
entry_column1.pack(side='left')
label_column2.pack(side='left')
entry_column2.pack(side='left')
label_column3.pack(side='left')
entry_column3.pack(side='left')
label_column4.pack(side='left')
entry_column4.pack(side='left')
label_column5.pack(side='left')
entry_column5.pack(side='left')

# 엔트리 위젯간의 탭 이동 설정
entry_column1.focus()
entry_column1.bind("<Return>", focus_next_entry)
entry_column2.bind("<Return>", focus_next_entry)
entry_column3.bind("<Return>", focus_next_entry)
entry_column4.bind("<Return>", focus_next_entry)
entry_column5.bind("<Return>", focus_next_entry)

# 마지막 엔트리에서 엔터를 누르면 저장되도록 설정
entry_column5.bind("<Return>", insert_data)

# 프레임 2: 버튼 프레임
button_frame = tk.Frame(app)
button_frame.pack(pady=5)

# 저장 버튼 (우측에 배치)
save_button = tk.Button(button_frame, text="저장", command=insert_data)
save_button.pack(side='right', padx=10)

# 삭제 버튼
delete_button = tk.Button(button_frame, text="삭제", command=delete_selected_row)
delete_button.pack(side='right', padx=10)

# Excel 파일로 데이터 추가 버튼
add_excel_button = tk.Button(button_frame, text="Excel 가져오기", command=insert_data_from_excel)
add_excel_button.pack(side='right', padx=10)

# Excel 파일로 데이터 내보내기 버튼
export_excel_button = tk.Button(button_frame, text="Excel 내보내기", command=export_data_to_excel)
export_excel_button.pack(side='right', padx=10)

# 프레임 3: Treeview
treeview_frame = tk.Frame(app)
treeview_frame.pack(pady=5)

# Treeview 폰트 설정
tree_style = ttk.Style()
tree_style.configure("Treeview", font=("Consolas", 10))

# Treeview 열 정렬 설정 (내선번호 열 기본 정렬은 오름차순)
tree = ttk.Treeview(treeview_frame, columns=('id', '기계선번', '내선번호', '수평선번', '사용자명', '기타입력'), show='headings', height=25, style="Treeview")

# 초기 정렬을 "내선번호" 열 오름차순으로 설정
current_sort_column = "내선번호"
ascending = True

# 열 헤더 클릭 시 정렬 함수 (내선번호 열 내림차순으로 정렬)
def sort_treeview_column(column):
    global ascending, current_sort_column

    if current_sort_column == column:
        ascending = not ascending
    else:
        ascending = True
        current_sort_column = column

    current_items = tree.get_children('')
    
    # 숫자로 변환 가능한 값만 정렬에 포함
    def sort_key(item):
        value = tree.set(item, column)
        try:
            return int(value)
        except ValueError:
            return value
    
    sorted_items = sorted(current_items, key=sort_key, reverse=ascending)  # ascending 값을 반대로 설정
    
    for index, item in enumerate(sorted_items):
        tree.move(item, '', index)
    
    # 정렬 방식 변경 시 아이콘 업데이트
    for col in tree['columns']:
        if col == column:
            tree.heading(col, text=col + (" ▼" if ascending else " ▲"), command=lambda c=col: sort_treeview_column(c))
        else:
            tree.heading(col, text=col, command=lambda c=col: sort_treeview_column(c))

# id 를 숨김
tree.column('id', stretch=False, minwidth=0, width=0)
# 수직 스크롤바 추가
v_scrollbar = ttk.Scrollbar(treeview_frame, orient=tk.VERTICAL, command=tree.yview)
v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# 수평 스크롤바 추가
h_scrollbar = ttk.Scrollbar(treeview_frame, orient=tk.HORIZONTAL, command=tree.xview)
h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

# Treeview 위젯과 스크롤바 연결
tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
tree.pack()


# 기본 정렬을 "내선번호" 열 오름차순으로 설정 (여기에서 호출)
sort_treeview_column("내선번호")

# Treeview 업데이트
update_treeview()

# F2를 눌러 수정하기
def f2_key(event):
    item = tree.selection()[0]
    open_edit_popup(item)
tree.bind("<F2>", f2_key)

# F5를 눌러 새로고침
def f5_key(event):
    update_treeview()
    messagebox.showinfo("새로고침", "새로고침")
tree.bind("<F5>", f5_key)

# 삭제 함수를 바인딩 = Delete 키
app.bind("<Delete>", delete_selected_row)
# 삭제 함수를 바인딩 = Delete 버튼
delete_button.config(command=delete_selected_row)

# 프레임 3: description_frame
description_frame = tk.Frame(app)
description_frame.pack(padx=20, anchor=tk.W)

# 설명 레이블 생성
description_label = tk.Label(description_frame, 
text='''1. 수평선번에 IP주소를 입력하세요. (단, 아날로그 전화인 경우 수평선번을 입력하세요.)

2. 수정할 항목을 선택하고 [F2]를 눌러 수정하세요.

3. 대량의 자료 입력의 경우 xlsx 파일을 사용하면 쉽게 수정할 수 있습니다.''')
description_label.pack(side=tk.BOTTOM, pady=5, anchor=tk.W)
# 텍스트 정렬
description_label.config(justify=tk.LEFT)

app.mainloop()
